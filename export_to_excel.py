import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from num2words import num2words

# Function to convert number to Arabic words
def number_to_arabic_words(number):
    # Convert the number to words in Arabic directly
    return num2words(number, lang='ar')

# Initialize a new Workbook
wb = Workbook()
ws = wb.active
ws.title = "Numbers to Arabic"

# Set headers
ws['A1'] = 'Number'
ws['B1'] = 'Arabic Words'

# Apply text alignment for the Arabic Words column
for cell in ws['B']:
    cell.alignment = Alignment(horizontal='right')

# Fill in the numbers and their corresponding Arabic words
for num in range(1, 1001):
    arabic_word = number_to_arabic_words(num)
    ws.append([num, arabic_word])

# Save the workbook
excel_filename = r'D:\numbers exports\numbers_to_arabic_words.xlsx'
wb.save(filename=excel_filename)

print(f"Exported to {excel_filename} successfully!")
