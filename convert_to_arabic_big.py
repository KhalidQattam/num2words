import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from num2words import num2words
import time

# Function to convert number to Arabic words
def number_to_arabic_words(number):
    return num2words(number, lang='ar')

# Function to save numbers and their Arabic words to an Excel file
def save_to_excel(start_num, end_num, folder_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Numbers to Arabic"
    ws['A1'] = 'Number'
    ws['B1'] = 'Arabic Words'
    for cell in ws['B']:
        cell.alignment = Alignment(horizontal='right')
    for num in range(start_num, end_num + 1):
        arabic_word = number_to_arabic_words(num)
        ws.append([num, arabic_word])
    excel_filename = f'{folder_path}arabic_numbers_{start_num}_{end_num}.xlsx'
    wb.save(filename=excel_filename)
    print(f"Exported to {excel_filename} successfully!")

folder_path = 'D:\\numbers exports\\'
range_start = 1000000
range_end = 10000000
step = 1000000

# Calculate total iterations for progress tracking
total_iterations = (range_end - range_start) // step
completed_iterations = 0

# Start timing the operation
start_time = time.time()

for start_num in range(range_start, range_end, step):
    end_num = start_num + step - 1
    save_to_excel(start_num, end_num, folder_path)
    completed_iterations += 1
    current_time = time.time()
    elapsed_time = current_time - start_time
    estimated_total_time = (elapsed_time / completed_iterations) * total_iterations
    estimated_time_remaining = estimated_total_time - elapsed_time
    completion_percentage = (completed_iterations / total_iterations) * 100
    print(f"Completion: {completion_percentage:.2f}%. Estimated time remaining: {estimated_time_remaining / 60:.2f} minutes.")

print("All files have been exported successfully.")
